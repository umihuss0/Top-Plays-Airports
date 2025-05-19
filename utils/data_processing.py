# utils/data_processing.py
import io
from datetime import datetime
from typing import Dict, Any

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ──────────────────────────────────────────────────────────────────────────────
#  Helper imports for Airport → Market tagging
# ──────────────────────────────────────────────────────────────────────────────
from components.market_drilldown import extract_airport, make_market_label


# ──────────────────────────────────────────────────────────────────────────────
#  MAIN FILE-PROCESSING FUNCTION
# ──────────────────────────────────────────────────────────────────────────────
@st.cache_resource
def process_file(uploaded_file) -> Dict[str, Any]:
    """
    Parse the uploaded Excel/CSV, build all required summaries,
    generate an Excel report, and return everything in a single dict.
    """
    # 1. ── LOAD ────────────────────────────────────────────────────────────
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    original_cols = df.columns.tolist()

    # 2. ── INITIAL CLEAN / EXTRA COLUMNS ───────────────────────────────────
    def parse_system(row):
        val = str(row["System"]).strip().upper()
        if "SPOTCHART" in val:
            return "Roadside"
        if any(x in val for x in ("ADPORTAL", "RTB ADSERVER", "VISTAR SCHEDULING SERVICE")):
            return "Airport"
        return "Unknown"

    def market_code(row):
        display = str(row["Display"]).upper()
        system_type = row["System_Type"]
        if system_type == "Roadside":
            import re

            m = re.match(r"\D*([A-Z]{3})", re.sub(r"^\d+", "", display))
            return m.group(1) if m else ""
        return ""

    def airport_group(row):
        system_type = row["System_Type"]
        network_code = row["Network Code"]
        if system_type == "Airport" and pd.notna(network_code):
            return str(network_code).split("_")[0].upper()
        return ""

    def hour_24(dt):
        if pd.isna(dt):
            return np.nan
        try:
            return pd.to_datetime(dt).hour
        except Exception:
            return np.nan

    def network_name(row):
        display = row["Display"]
        system_type = row["System_Type"]
        return str(display).upper() if system_type == "Airport" else ""

    # ------ apply cleaning helpers ----------------------------------------
    df["System_Type"] = df.apply(parse_system, axis=1)
    df["Hour_24"] = df["Date & Hour - EST"].apply(hour_24)
    df["Market_Code"] = df.apply(market_code, axis=1)
    df["Airport_Group"] = df.apply(airport_group, axis=1)
    df["Network_Name"] = df.apply(network_name, axis=1)
    df["# Plays"] = pd.to_numeric(df["# Plays"], errors="coerce").fillna(0).astype("int64")

    # 3. ── ***NEW***  TAG AIRPORT + MARKET FOR DRILL-DOWN ──────────────────
    # normalise header
    if "Network Code" in df.columns and "Network_Code" not in df.columns:
        df = df.rename(columns={"Network Code": "Network_Code"})

    if "Network_Code" not in df.columns:
        raise ValueError("Column 'Network Code' not found in the uploaded file")

    df["Airport"] = df["Network_Code"].apply(extract_airport)
    df["Market"] = df["Airport"].apply(make_market_label)

    # 4. ── BUILD SUMMARY TABLES ────────────────────────────────────────────
    raw_data = df.copy()

    # overall hourly
    overall_hourly = (
        df.groupby("Hour_24", as_index=False)["# Plays"]
        .sum()
        .rename(columns={"# Plays": "Total_Plays"})
    )
    overall_hourly = overall_hourly.sort_values(
        ["Total_Plays", "Hour_24"], ascending=[False, True]
    )
    overall_hourly["Rank"] = (
        overall_hourly.groupby("Total_Plays", group_keys=False).cumcount() + 1
    )
    overall_hourly["Rank"] = (
        overall_hourly["Rank"]
        + overall_hourly["Total_Plays"].rank(method="min", ascending=False).astype(int)
        - 1
    )
    overall_hourly = overall_hourly.sort_values(["Rank", "Hour_24"])

    # ➜ Roadside / Airport summaries (unchanged from your original code)
    roadside = df[df["System_Type"] == "Roadside"]
    roadside_hourly = (
        roadside.groupby("Hour_24", as_index=False)["# Plays"]
        .sum()
        .rename(columns={"# Plays": "Total_Plays"})
    )
    roadside_hourly = roadside_hourly.sort_values(
        ["Total_Plays", "Hour_24"], ascending=[False, True]
    )
    roadside_hourly["Rank"] = (
        roadside_hourly.groupby("Total_Plays", group_keys=False).cumcount() + 1
    )
    roadside_hourly["Rank"] = (
        roadside_hourly["Rank"]
        + roadside_hourly["Total_Plays"]
        .rank(method="min", ascending=False)
        .astype(int)
        - 1
    )
    roadside_hourly = roadside_hourly.sort_values(["Rank", "Hour_24"])

    roadside_by_market = (
        roadside.groupby(["Market_Code", "Hour_24"], as_index=False)["# Plays"]
        .sum()
        .rename(columns={"# Plays": "Market_Plays"})
    )
    roadside_by_market = roadside_by_market.sort_values(
        ["Market_Code", "Market_Plays", "Hour_24"], ascending=[True, False, True]
    )
    roadside_by_market["Rank_within_Market"] = roadside_by_market.groupby(
        "Market_Code"
    )["Market_Plays"].rank(method="min", ascending=False).astype(int)

    airport = df[df["System_Type"] == "Airport"]
    airport_hourly = (
        airport.groupby("Hour_24", as_index=False)["# Plays"]
        .sum()
        .rename(columns={"# Plays": "Total_Plays"})
    )
    airport_hourly = airport_hourly.sort_values(
        ["Total_Plays", "Hour_24"], ascending=[False, True]
    )
    airport_hourly["Rank"] = (
        airport_hourly.groupby("Total_Plays", group_keys=False).cumcount() + 1
    )
    airport_hourly["Rank"] = (
        airport_hourly["Rank"]
        + airport_hourly["Total_Plays"].rank(method="min", ascending=False).astype(int)
        - 1
    )
    airport_hourly = airport_hourly.sort_values(["Rank", "Hour_24"])

    airport_by_group = (
        airport.groupby(["Airport_Group", "Hour_24"], as_index=False)["# Plays"]
        .sum()
        .rename(columns={"# Plays": "Group_Plays"})
    )
    airport_by_group = airport_by_group.sort_values(
        ["Airport_Group", "Group_Plays", "Hour_24"], ascending=[True, False, True]
    )
    airport_by_group["Rank_within_Group"] = airport_by_group.groupby(
        "Airport_Group"
    )["Group_Plays"].rank(method="min", ascending=False).astype(int)

    airport_by_network = (
        airport.groupby(["Network_Name", "Hour_24"], as_index=False)["# Plays"]
        .sum()
        .rename(columns={"# Plays": "Network_Plays"})
    )
    airport_by_network = airport_by_network.sort_values(
        ["Network_Name", "Network_Plays", "Hour_24"], ascending=[True, False, True]
    )
    airport_by_network["Rank_within_Network"] = airport_by_network.groupby(
        "Network_Name"
    )["Network_Plays"].rank(method="min", ascending=False).astype(int)

    # 5. ── CREATE EXCEL REPORT (unchanged) ─────────────────────────────────
    output = io.BytesIO()
    wb = openpyxl.Workbook()

    ws_raw = wb.active
    ws_raw.title = "Raw_Data"
    for r in dataframe_to_rows(raw_data, index=False, header=True):
        ws_raw.append(r)

    ws_top = wb.create_sheet("Top_Hours_Overall")
    for r in dataframe_to_rows(overall_hourly.head(10), index=False, header=True):
        ws_top.append(r)

    ws_rs_sum = wb.create_sheet("Roadside_Summary")
    for r in dataframe_to_rows(roadside_hourly.head(5), index=False, header=True):
        ws_rs_sum.append(r)

    ws_rs_market = wb.create_sheet("Roadside_By_Market")
    for r in dataframe_to_rows(roadside_by_market, index=False, header=True):
        ws_rs_market.append(r)

    ws_ap_sum = wb.create_sheet("Airport_Summary")
    for r in dataframe_to_rows(airport_hourly.head(5), index=False, header=True):
        ws_ap_sum.append(r)

    ws_ap_group = wb.create_sheet("Airport_By_Group")
    for r in dataframe_to_rows(airport_by_group, index=False, header=True):
        ws_ap_group.append(r)

    ws_ap_net = wb.create_sheet("Airport_By_Network")
    for r in dataframe_to_rows(airport_by_network, index=False, header=True):
        ws_ap_net.append(r)

    # basic header formatting / highlighting (same as your original)
    bold = Font(bold=True)
    fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = bold
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, int):
                    cell.number_format = "#,##0"
        if "Rank" in [c.value for c in ws[1]]:
            rank_idx = [c.value for c in ws[1]].index("Rank") + 1
            for row in ws.iter_rows(min_row=2):
                if row[rank_idx - 1].value == 1:
                    for cell in row:
                        cell.fill = fill

        # similar highlight logic for other rank columns …
        for special in (
            "Rank_within_Market",
            "Rank_within_Group",
            "Rank_within_Network",
        ):
            if special in [c.value for c in ws[1]]:
                idx = [c.value for c in ws[1]].index(special) + 1
                for row in ws.iter_rows(min_row=2):
                    if row[idx - 1].value == 1:
                        for cell in row:
                            cell.fill = fill

    wb.save(output)
    output.seek(0)

    # 6. ── PACKAGE & RETURN ────────────────────────────────────────────────
    summary: Dict[str, Any] = {
        "overall_hourly": overall_hourly,
        "roadside_hourly": roadside_hourly,
        "roadside_by_market": roadside_by_market,
        "airport_hourly": airport_hourly,
        "airport_by_group": airport_by_group,
        "airport_by_network": airport_by_network,
        "raw": raw_data,  #  ← used by Market Drill-down
        "report_bytes": output.read(),
    }
    return summary
