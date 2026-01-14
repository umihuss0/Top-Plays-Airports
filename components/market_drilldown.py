import streamlit as st
import pandas as pd
from io import BytesIO
# import numpy as np # Not strictly needed if using .mean() on Series and handling empty slices

# ──────────────────────────────────────────────────────────────────────────────
# Constants for Prime Play Window logic
# ──────────────────────────────────────────────────────────────────────────────
MIN_PRIME_WINDOW_LEN = 2
MAX_PRIME_WINDOW_LEN = 4
MIN_PLAYS_PERCENT_OF_RANGE1 = 0.70 # For justifying a second window

# Constants for new scoring logic (as per "Recommended scoring tweaks")
LOW_HOUR_RELATIVE_THRESHOLD = 0.70  # Min hour < 70% of window's *density* (avg plays/hr)
WEAK_HOUR_PENALTY_PER_FRACTION = 0.20 # -20% to density score for each weak hour *fraction* of window


# ─── AIRPORT DICTIONARIES ────
AIRPORT_TO_MARKET = {
    "ATL": "Atlanta", "AUS": "Austin", "BNA": "Nashville", "BTR": "Baton Rouge",
    "BWI": "Baltimore / Washington", "CAK": "Akron / Canton", "CLE": "Cleveland",
    "CMH": "Columbus", "CRP": "Corpus Christi", "DAB": "Daytona Beach",
    "DCA": "Washington DC", "DEN": "Denver", "DTW": "Detroit", "ELP": "El Paso",
    "EWR": "New Jersey", "FAR": "Fargo", "FAT": "Fresno", "FLL": "Ft. Lauderdale",
    "GSO": "Greensboro", "HNL": "Honolulu", "HSV": "Huntsville", "IAD": "Washington DC",
    "ICT": "Wichita", "JAN": "Jackson (MS)", "JFK": "New York", "LBB": "Lubbock",
    "LGA": "New York", "MDT": "Harrisburg", "MDW": "Chicago",
    "MSP": "Minneapolis / St Paul", "MSY": "New Orleans", "OMA": "Omaha",
    "ORD": "Chicago", "PBI": "Palm Beach", "PHL": "Philadelphia", "PSC": "Tri-Cities",
    "RDU": "Raleigh-Durham", "ROA": "Roanoke–Blacksburg", "SAN": "San Diego",
    "SEA": "Seattle", "SFO": "San Francisco", "SMF": "Sacramento",
    "SWF": "Hudson Valley / NY", "TLH": "Tallahassee", "VPS": "Destin-Fort Walton",  "SFS": "FBO",
}

AIRPORT_LONG_NAME = {
    "ATL": "Hartsfield-Jackson Atlanta International Airport",
    "AUS": "Austin-Bergstrom International Airport", "BNA": "Nashville International Airport",
    "BTR": "Baton Rouge Metropolitan Airport",
    "BWI": "Baltimore/Washington International Thurgood Marshall Airport",
    "CAK": "Akron-Canton Airport", "CLE": "Cleveland-Hopkins International Airport",
    "CMH": "John Glenn Columbus International Airport",
    "CRP": "Corpus Christi International Airport",
    "DAB": "Daytona Beach International Airport",
    "DCA": "Ronald Reagan Washington National Airport", "DEN": "Denver International Airport",
    "DTW": "Detroit Metropolitan Wayne County Airport", "ELP": "El Paso International Airport",
    "EWR": "Newark Liberty International Airport", "FAR": "Hector International Airport",
    "FAT": "Fresno Yosemite International Airport",
    "FLL": "Ft. Lauderdale-Hollywood International Airport",
    "GSO": "Piedmont Triad International Airport",
    "HNL": "Daniel K. Inouye International Airport", "HSV": "Huntsville International Airport",
    "IAD": "Washington Dulles International Airport",
    "ICT": "Wichita Dwight D. Eisenhower National Airport",
    "JAN": "Jackson-Medgar Wiley Evers International Airport",
    "JFK": "John F. Kennedy International Airport",
    "LBB": "Lubbock Preston Smith International Airport", "LGA": "LaGuardia Airport",
    "MDT": "Harrisburg International Airport", "MDW": "Chicago Midway International Airport",
    "MSP": "Minneapolis-St. Paul International Airport",
    "MSY": "Louis Armstrong New Orleans International Airport",
    "OMA": "Omaha Eppley Airfield", "ORD": "Chicago O’Hare International Airport",
    "PBI": "Palm Beach International Airport", "PHL": "Philadelphia International Airport",
    "PSC": "Tri-Cities Airport", "RDU": "Raleigh-Durham International Airport",
    "ROA": "Roanoke-Blacksburg Regional Airport", "SAN": "San Diego International Airport",
    "SEA": "Seattle-Tacoma International Airport", "SFO": "San Francisco International Airport",
    "SMF": "Sacramento International Airport", "SWF": "New York Stewart International Airport",
    "TLH": "Tallahassee International Airport", "VPS": "Destin-Fort Walton Beach Airport", "SFS": "Various Private Aviation Terminals",
}
# ──────────────────────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────────────────────
# Helper Functions
# ──────────────────────────────────────────────────────────────────────────────
def extract_airport(network_code: str | float) -> str | None:
    if pd.isna(network_code): return None
    if isinstance(network_code, str):
        stripped_full_code = network_code.strip()
        if stripped_full_code:
            parts = stripped_full_code.split("_")
            if parts: return parts[0].strip().upper()
    return None

def make_market_label(airport_code: str | None) -> str:
    if not airport_code: return "Roadside Markets"
    code_to_lookup = str(airport_code).strip().upper()
    market = AIRPORT_TO_MARKET.get(code_to_lookup)
    if market:
        return market
    # If not a recognized airport code, treat as roadside
    return "Roadside Markets"

def format_hour(hour_24: int | float | None) -> str:
    if pd.isna(hour_24): return "-"
    h = int(hour_24)
    if h in (0, 24): return "12am"
    if h == 12: return "12pm"
    return f"{h}am" if h < 12 else f"{h-12}pm"

def vertical_spacer(height_px: int = 24) -> None: # Renamed arg to avoid conflict with px module
    st.markdown(f"<div style='height:{height_px}px'></div>", unsafe_allow_html=True)

def format_currency(amount: float | int | None, rounded: bool = False) -> str:
    """Format a number as currency with $ sign and comma separators.

    Args:
        amount: The amount to format
        rounded: If True, round to whole dollars (no decimals) for collapsed views
    """
    if pd.isna(amount) or amount is None:
        return "$0"
    if rounded:
        return f"${int(round(amount)):,}"
    return f"${amount:,.2f}" if isinstance(amount, float) else f"${int(amount):,}"

def get_revenue_column(df: pd.DataFrame) -> str | None:
    """Find the revenue column in the dataframe (handles various naming conventions)."""
    possible_names = ["Revenue", "revenue", "REVENUE", "Revenue ($)", "revenue ($)", "Total Revenue"]
    for name in possible_names:
        if name in df.columns:
            return name
    return None

def create_revenue_excel(df: pd.DataFrame, revenue_col: str, sorted_markets: list,
                         advertiser_name: str = "", date_range: str = "") -> bytes:
    """Create a formatted Excel file with revenue breakdown."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = []
        for market_name in sorted_markets:
            market_df = df[df["Market"] == market_name]
            market_total = market_df[revenue_col].sum()
            if market_total > 0:
                summary_data.append({
                    "Market": market_name,
                    "Revenue": market_total
                })

        summary_df = pd.DataFrame(summary_data)
        if not summary_df.empty:
            # Add total row
            total_row = pd.DataFrame([{"Market": "TOTAL", "Revenue": summary_df["Revenue"].sum()}])
            summary_df = pd.concat([summary_df, total_row], ignore_index=True)

        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Detailed breakdown sheet
        detail_data = []
        for market_name in sorted_markets:
            market_df = df[df["Market"] == market_name]
            network_revenue = (
                market_df.groupby("Network_Code", as_index=False)[revenue_col]
                .sum()
                .sort_values(revenue_col, ascending=False)
            )
            for _, row in network_revenue.iterrows():
                if row[revenue_col] > 0:
                    detail_data.append({
                        "Market": market_name,
                        "Network Code": row["Network_Code"],
                        "Revenue": row[revenue_col]
                    })

        detail_df = pd.DataFrame(detail_data)
        if not detail_df.empty:
            detail_df.to_excel(writer, sheet_name="Detailed Breakdown", index=False)

        # Format the sheets
        workbook = writer.book
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            # Set column widths
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 25
            if sheet_name == "Detailed Breakdown":
                worksheet.column_dimensions['C'].width = 15

            # Format revenue column as currency
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            # Header styling
            header_fill = PatternFill(start_color="4F8EF7", end_color="4F8EF7", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )

            for col_num, cell in enumerate(worksheet[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Roadside highlight (light red/pink)
            roadside_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            roadside_font = Font(italic=True)

            # Format data rows
            for row in worksheet.iter_rows(min_row=2):
                is_roadside = False
                # Check if this row contains "Roadside Markets"
                for cell in row:
                    if cell.value == "Roadside Markets":
                        is_roadside = True
                        break

                for cell in row:
                    cell.border = thin_border
                    if cell.column_letter in ['B', 'C'] and isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
                    # Apply roadside styling
                    if is_roadside:
                        cell.fill = roadside_fill
                        if cell.column_letter == 'A':
                            cell.font = roadside_font

    output.seek(0)
    return output.getvalue()

# ──────────────────────────────────────────────────────────────────────────────
# Prime Play Window Finder Logic (Incorporating new scoring and edge trimming)
# ──────────────────────────────────────────────────────────────────────────────
# In components/market_drilldown.py
# ... (constants and other helper functions remain the same) ...

# ──────────────────────────────────────────────────────────────────────────────
# Prime Play Window Finder Logic (Updated for Window #2 Trimming Debug)
# ──────────────────────────────────────────────────────────────────────────────
def find_prime_play_windows(overall_hourly_df: pd.DataFrame | None) -> list:
    # ... (Initial checks and df_filtered preparation - same as before) ...
    if overall_hourly_df is None or overall_hourly_df.empty: return []
    if 'Hour_24' not in overall_hourly_df.columns or 'Total_Plays' not in overall_hourly_df.columns: return []

    df = overall_hourly_df.copy()
    df['Hour_24'] = pd.to_numeric(df['Hour_24'], errors='coerce')
    df.dropna(subset=['Hour_24'], inplace=True)
    df['Hour_24'] = df['Hour_24'].astype(int)
    df['Total_Plays'] = pd.to_numeric(df['Total_Plays'], errors='coerce').fillna(0)
    df_filtered = df.query("7 <= Hour_24 <= 21").sort_values("Hour_24").reset_index(drop=True)
    if df_filtered.empty: return []
    
    all_windows = []
    n_hours_available = len(df_filtered)
    if n_hours_available < MIN_PRIME_WINDOW_LEN: return []

    # ... (window generation loop - same as before) ...
    for i in range(n_hours_available):
        start_hour_series = df_filtered['Hour_24'].iloc[i]
        for dur in range(MIN_PRIME_WINDOW_LEN, MAX_PRIME_WINDOW_LEN + 1):
            if i + dur > n_hours_available: break
            slice_df = df_filtered.iloc[i : i + dur]
            if slice_df['Hour_24'].iloc[-1] != start_hour_series + dur - 1: break

            total_plays = slice_df['Total_Plays'].sum()
            density = total_plays / dur if dur > 0 else 0
            weak_mask = slice_df['Total_Plays'] < LOW_HOUR_RELATIVE_THRESHOLD * density
            weak_count = weak_mask.sum()
            penalty_multiplier = 1.0 - (WEAK_HOUR_PENALTY_PER_FRACTION * (weak_count / dur)) if dur > 0 else 1.0
            rank_score = density * penalty_multiplier

            all_windows.append({
                "Window_Start": int(slice_df['Hour_24'].iloc[0]),
                "Window_End": int(slice_df['Hour_24'].iloc[-1]),
                "Window_Duration": int(dur),
                "Window_Total_Plays": int(total_plays),
                "Density": round(density, 2),
                "Weak_Count": int(weak_count),
                "Rank_Score": round(rank_score, 2)
            })
    
    if not all_windows: return []
    windows_df = pd.DataFrame(all_windows)
    if windows_df.empty: return []

    windows_df_sorted = windows_df.sort_values(
        by=["Rank_Score", "Window_Duration", "Window_Start"],
        ascending=[False, True, True]
    ).reset_index(drop=True)
    
    if windows_df_sorted.empty: return []

    # --- Helper function for trimming a window ---
    def _trim_window(start_hr, end_hr, duration_val, source_df_filtered): # Removed context_debug
        _current_start = int(start_hr)
        _current_end = int(end_hr)
        _current_duration = int(duration_val)

        # Trim weak start hour
        while _current_duration > MIN_PRIME_WINDOW_LEN:
            first_hour_plays_series = source_df_filtered[source_df_filtered['Hour_24'] == _current_start]['Total_Plays']
            if first_hour_plays_series.empty: 
                break
            first_hour_plays = first_hour_plays_series.iloc[0]

            current_window_segment_df = source_df_filtered[
                (source_df_filtered['Hour_24'] >= _current_start) & (source_df_filtered['Hour_24'] <= _current_end)
            ]
            if current_window_segment_df.empty or _current_duration == 0: 
                break 
            current_segment_total_plays = current_window_segment_df['Total_Plays'].sum()
            current_segment_density = current_segment_total_plays / _current_duration

            if first_hour_plays < LOW_HOUR_RELATIVE_THRESHOLD * current_segment_density:
                _current_start += 1
                _current_duration -= 1
            else:
                break 

        # Trim weak end hour
        while _current_duration > MIN_PRIME_WINDOW_LEN:
            last_hour_plays_series = source_df_filtered[source_df_filtered['Hour_24'] == _current_end]['Total_Plays']
            if last_hour_plays_series.empty: 
                break
            last_hour_plays = last_hour_plays_series.iloc[0]
            
            current_window_segment_df = source_df_filtered[
                (source_df_filtered['Hour_24'] >= _current_start) & (source_df_filtered['Hour_24'] <= _current_end)
            ]
            if current_window_segment_df.empty or _current_duration == 0: 
                break
            current_segment_total_plays = current_window_segment_df['Total_Plays'].sum()
            current_segment_density = current_segment_total_plays / _current_duration
            
            if last_hour_plays < LOW_HOUR_RELATIVE_THRESHOLD * current_segment_density:
                _current_end -= 1
                _current_duration -= 1
            else:
                break
        
        final_trimmed_plays_df = source_df_filtered[
            (source_df_filtered['Hour_24'] >= _current_start) & (source_df_filtered['Hour_24'] <= _current_end)
        ]
        final_trimmed_plays = final_trimmed_plays_df['Total_Plays'].sum() if not final_trimmed_plays_df.empty else 0
        
        return _current_start, _current_end, _current_duration, int(final_trimmed_plays)

    # --- Select and Trim Window #1 ---
    best_window1_original_series = windows_df_sorted.iloc[0]
    w1_trimmed_start, w1_trimmed_end, w1_trimmed_duration, w1_trimmed_plays = _trim_window(
        best_window1_original_series["Window_Start"],
        best_window1_original_series["Window_End"],
        best_window1_original_series["Window_Duration"],
        df_filtered
    )
    result = [(w1_trimmed_start, w1_trimmed_end, w1_trimmed_plays)]

    # --- Determine if a second window is justified (and trim it) ---
    range1_plays_for_comparison = w1_trimmed_plays 
    potential_range2_candidates = windows_df_sorted.iloc[1:].copy() 
    selected_window2_details = None

    for idx, candidate_w2_series in potential_range2_candidates.iterrows():
        no_overlap = (candidate_w2_series["Window_End"] < w1_trimmed_start) or \
                     (candidate_w2_series["Window_Start"] > w1_trimmed_end)
        if not no_overlap:
            continue

        initial_plays_ok = candidate_w2_series["Window_Total_Plays"] >= MIN_PLAYS_PERCENT_OF_RANGE1 * range1_plays_for_comparison
        if not initial_plays_ok:
            continue
        
        # If initial checks pass, trim this candidate Window #2
        w2_cand_trimmed_start, w2_cand_trimmed_end, w2_cand_trimmed_duration, w2_cand_trimmed_plays = _trim_window(
            candidate_w2_series["Window_Start"],
            candidate_w2_series["Window_End"],
            candidate_w2_series["Window_Duration"],
            df_filtered
        )

        if w2_cand_trimmed_duration < MIN_PRIME_WINDOW_LEN:
             continue

        final_plays_ok = w2_cand_trimmed_plays >= MIN_PLAYS_PERCENT_OF_RANGE1 * range1_plays_for_comparison
        
        if final_plays_ok:
            selected_window2_details = (w2_cand_trimmed_start, w2_cand_trimmed_end, w2_cand_trimmed_plays)
            break 

    if selected_window2_details:
        result.append(selected_window2_details)
            
    return result
# ──────────────────────────────────────────────────────────────────────────────
# Main render function
# ──────────────────────────────────────────────────────────────────────────────
def render_market_drilldown() -> None:
    raw_df: pd.DataFrame | None = st.session_state.get("data", {}).get("raw")

    if raw_df is None or raw_df.empty:
        st.info("Upload a report to see market details.")
        return

    df = raw_df.copy()
    if "Network Code" in df.columns and "Network_Code" not in df.columns:
        df = df.rename(columns={"Network Code": "Network_Code"})

    if "Network_Code" not in df.columns:
        st.error("Column 'Network Code' not found in the upload.")
        return
    
    # Ensure 'Hour_24' column exists and is numeric for grouping
    if 'Hour_24' not in df.columns:
        st.error("Column 'Hour_24' not found in the upload, which is required for hourly analysis.")
        return
    df['Hour_24'] = pd.to_numeric(df['Hour_24'], errors='coerce')
    df.dropna(subset=['Hour_24'], inplace=True) # Remove rows where Hour_24 could not be converted
    df['Hour_24'] = df['Hour_24'].astype(int)

    # Ensure '# Plays' column exists and is numeric
    if '# Plays' not in df.columns:
        st.error("Column '# Plays' not found in the upload, which is required for hourly analysis.")
        return
    df['# Plays'] = pd.to_numeric(df['# Plays'], errors='coerce').fillna(0)


    if "Airport" not in df.columns:
        df["Airport"] = df["Network_Code"].apply(extract_airport)
    else:
        df["Airport"] = df["Airport"].apply(lambda x: str(x).strip().upper() if pd.notna(x) else None)

    df["Market"] = df["Airport"].apply(make_market_label)

    # ── CHECK FOR REVENUE COLUMN ──────────────────────────────────────────────
    revenue_col = get_revenue_column(df)
    has_revenue = revenue_col is not None

    if has_revenue:
        df[revenue_col] = pd.to_numeric(df[revenue_col], errors='coerce').fillna(0)

    st.write("") # Creates a bit of space before the first expander
    # Sort markets alphabetically, but always put "Roadside Markets" at the end
    all_markets = df["Market"].dropna().unique().tolist()
    sorted_market_labels = sorted([m for m in all_markets if m != "Roadside Markets"])
    if "Roadside Markets" in all_markets:
        sorted_market_labels.append("Roadside Markets")

    # ── ALL MARKET REVENUE SECTION (shown first, before individual markets) ────
    if has_revenue:
        total_revenue = df[revenue_col].sum()

        # Try to extract advertiser name from data (look for Advertiser column or use generic label)
        advertiser_col = None
        for col_name in ["Advertiser", "advertiser", "ADVERTISER", "Advertiser Name"]:
            if col_name in df.columns:
                advertiser_col = col_name
                break

        advertiser_name = ""
        if advertiser_col and not df[advertiser_col].dropna().empty:
            # Get the most common advertiser name
            advertiser_name = df[advertiser_col].mode().iloc[0] if not df[advertiser_col].mode().empty else ""

        # Get date range from data
        date_range_str = ""
        date_col = None
        for col_name in ["Date & Hour - EST", "Date", "date", "DATE", "Date & Hour"]:
            if col_name in df.columns:
                date_col = col_name
                break

        if date_col:
            try:
                dates = pd.to_datetime(df[date_col], errors='coerce').dropna()
                if not dates.empty:
                    min_date = dates.min().strftime("%m/%d/%Y")
                    max_date = dates.max().strftime("%m/%d/%Y")
                    date_range_str = f"{min_date} – {max_date}"
            except:
                pass

        revenue_header_label = f"{advertiser_name} All Market Revenue" if advertiser_name else "All Market Revenue"

        with st.expander(f"{revenue_header_label}   [{format_currency(total_revenue, rounded=True)}]", expanded=False):
            # Show date range if available
            if date_range_str:
                st.markdown(f'<p style="color: #666; font-size: 0.85rem; margin: 0 0 12px 0;"><em>Date Range: {date_range_str}</em></p>', unsafe_allow_html=True)

            # Build complete HTML for all markets
            all_markets_html = '<div class="revenue-breakdown-card">'

            # Group revenue by market and network code
            for market_name in sorted_market_labels:
                market_revenue_df = df[df["Market"] == market_name]
                market_total_revenue = market_revenue_df[revenue_col].sum()

                if market_total_revenue <= 0:
                    continue

                # Get airport code for this market (filter out None values)
                airport_codes_in_market = [c for c in market_revenue_df["Airport"].dropna().unique() if c and c in AIRPORT_TO_MARKET]
                airport_label = f"({', '.join(sorted(airport_codes_in_market))})" if airport_codes_in_market else ""

                # Highlight roadside markets with different styling
                is_roadside = market_name == "Roadside Markets"
                market_title_style = "font-style: italic;" if is_roadside else ""
                roadside_class = " roadside" if is_roadside else ""

                all_markets_html += f'''<div class="market-revenue-item{roadside_class}">
                    <div class="market-revenue-header">
                        <span class="market-revenue-title" style="{market_title_style}">{market_name} {airport_label}</span>
                        <span class="revenue-amount">{format_currency(market_total_revenue)}</span>
                    </div>'''

                # Breakdown by network code within this market
                network_revenue = (
                    market_revenue_df.groupby("Network_Code", as_index=False)[revenue_col]
                    .sum()
                    .sort_values(revenue_col, ascending=False)
                )

                for _, net_row in network_revenue.iterrows():
                    net_code = net_row["Network_Code"]
                    net_rev = net_row[revenue_col]
                    if net_rev > 0:
                        all_markets_html += f'''<div class="network-revenue-row">
                            <span class="network-code-label">{net_code}</span>
                            <span class="network-revenue-value">{format_currency(net_rev)}</span>
                        </div>'''

                all_markets_html += '</div>'

            # Total row at the bottom
            all_markets_html += f'''<div class="revenue-breakdown-row total-row">
                <span style="font-weight: 600;">Total Revenue</span>
                <span class="revenue-amount total">{format_currency(total_revenue)}</span>
            </div></div>'''

            st.markdown(all_markets_html, unsafe_allow_html=True)

            # Download button for Excel export
            vertical_spacer(12)
            excel_data = create_revenue_excel(df, revenue_col, sorted_market_labels, advertiser_name, date_range_str)
            file_name = f"{advertiser_name}_Revenue_Report.xlsx" if advertiser_name else "Revenue_Report.xlsx"
            file_name = file_name.replace(" ", "_")

            st.download_button(
                label="Download Revenue Report",
                data=excel_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary"
            )

        vertical_spacer(10)

    for market_group_name in sorted_market_labels:
        market_group_df = df[df["Market"] == market_group_name].copy()
        unique_airport_codes_in_group = sorted(market_group_df["Airport"].dropna().unique())

        for airport_code in unique_airport_codes_in_group:
            market_df = market_group_df[market_group_df["Airport"] == airport_code].copy()
            if market_df.empty: continue

            networks = sorted(market_df["Network_Code"].dropna().unique())
            market_name_from_dict = make_market_label(airport_code)
            long_name_from_dict = AIRPORT_LONG_NAME.get(str(airport_code) if airport_code else "", "")

            expander_main_label = f"{market_name_from_dict} ({airport_code})"
            if long_name_from_dict:
                expander_main_label += f" – {long_name_from_dict}"

            # Add revenue to the expander label if available (rounded for collapsed view)
            if has_revenue:
                market_revenue = market_df[revenue_col].sum()
                expander_main_label += f"   [{format_currency(market_revenue, rounded=True)}]"

            network_display_items = [f"[{n}]" for n in networks]
            second_line_html = f'<span class="network-label">Networks used:</span> {", ".join(network_display_items) if network_display_items else "–"}'

            with st.expander(expander_main_label, expanded=False):
                st.markdown(second_line_html, unsafe_allow_html=True)
                
                hourly = (
                    market_df.groupby("Hour_24", as_index=False)["# Plays"]
                    .sum()
                    .rename(columns={"# Plays": "Total_Plays"}) # This provides Total_Plays
                )
                # Ensure Total_Plays is numeric here as well, though '# Plays' conversion earlier should handle it.
                hourly["Total_Plays"] = pd.to_numeric(hourly["Total_Plays"], errors='coerce').fillna(0)
                hourly = hourly.sort_values(["Total_Plays", "Hour_24"], ascending=[False, True])

                if not hourly.empty:
                    hourly["Rank"] = hourly["Total_Plays"].rank(method="dense", ascending=False).astype(int)
                    hourly = hourly.sort_values(["Rank", "Hour_24"])
                else:
                    # Ensure `hourly` has the right columns even if empty for `find_prime_play_windows`
                    hourly = pd.DataFrame(columns=["Hour_24", "Total_Plays", "Rank"])


                # ── PRIME PLAY WINDOW(S) SECTION FOR THIS AIRPORT (NEW) ──────────────
                # Using h4 or a bolded st.markdown for subsection title
                st.markdown('**Prime Play Windows**', unsafe_allow_html=False) # Simple bold text
                # or st.markdown('<h4 class="subsection-title">Prime Play Window(s)</h4>', unsafe_allow_html=True)

                prime_windows_list_airport = find_prime_play_windows(hourly) # Pass the airport-specific hourly data

                if not prime_windows_list_airport:
                    st.markdown(
                        """
                        <div class="kpi-card" style="text-align: center; padding: 10px 0; margin-bottom: 10px;">
                            <span style="color: #6c757d; font-size: 0.85rem;">No qualifying hours (7 am – 9 pm)</span>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                else:
                    # Using st.columns for horizontal layout of prime window cards.
                    prime_window_cols = st.columns(len(prime_windows_list_airport))

                    for idx, window_data in enumerate(prime_windows_list_airport):
                        start_hour, end_hour, total_plays = window_data
                        
                        range_str = f"{format_hour(start_hour)} – {format_hour(end_hour)}"
                        plays_str = f"{total_plays:,} Total Plays"

                        card_label = "Prime Window"
                        if len(prime_windows_list_airport) > 1:
                            card_label = f"Window {idx + 1}"
                        
                        with prime_window_cols[idx]:
                            # You can add a class like 'range-card' for specific styling
                            prime_window_cols[idx].markdown(
                                f"""
                                <div class="kpi-card range-card">
                                    <div class="summary-title">{card_label}</div>
                                    <div class="summary-value">{range_str}</div>
                                    <div class="summary-subtitle">{plays_str}</div>
                                </div>
                                """,
                                unsafe_allow_html=True,
                            )
                vertical_spacer(15) # Space before Top Hour KPIs

                # --- Existing Top 3 KPI cards ---
                top3 = hourly.head(3)
                titles = ["Top Hour", "2nd Best", "3rd Best"]
                cols = st.columns(3)
                for idx_kpi, col_kpi in enumerate(cols):
                    with col_kpi:
                        if idx_kpi < len(top3):
                            row = top3.iloc[idx_kpi]
                            hour_val, total_plays_val = row.get("Hour_24"), row.get("Total_Plays")
                            hour_display = format_hour(hour_val)
                            plays_display = f"{int(total_plays_val):,} Total Plays" if pd.notna(total_plays_val) and total_plays_val > 0 else ("0 Total Plays" if pd.notna(total_plays_val) else "No Data")
                        else:
                            hour_display, plays_display = "-", "No Data"
                        col_kpi.markdown(f'<div class="kpi-card"><div class="summary-title">{titles[idx_kpi]}</div><div class="summary-value">{hour_display}</div><div class="summary-subtitle">{plays_display}</div></div>', unsafe_allow_html=True)
                
                vertical_spacer() # Existing spacer
                st.markdown('<h3 class="section-title">Top 10 Hours by Total Plays</h3>', unsafe_allow_html=True)

                if not hourly.empty:
                    tbl_data = hourly.head(10).copy()
                    tbl_data["Hour"] = tbl_data["Hour_24"].apply(format_hour)
                    # Rank should already be there from earlier calculation
                    if "Rank" not in tbl_data.columns: 
                        tbl_data["Rank"] = tbl_data["Total_Plays"].rank(method="dense", ascending=False).astype(int)
                    
                    tbl_data = tbl_data[["Hour", "Total_Plays", "Rank"]].rename(columns={"Total_Plays": "Total Plays"})
                    tbl_data["Total Plays"] = pd.to_numeric(tbl_data["Total Plays"], errors='coerce').fillna(0).astype(int)
                    tbl_data["Rank"] = pd.to_numeric(tbl_data["Rank"], errors='coerce').fillna(0).astype(int)
                    st.dataframe(tbl_data.style.format({"Total Plays": "{:,}", "Rank": "{}"}), use_container_width=True, hide_index=True)
                else:
                    st.caption("No hourly data to display for this airport.")